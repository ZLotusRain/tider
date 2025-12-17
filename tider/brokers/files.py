import os
import re
import csv
import json
import xlrd
import zipfile
import openpyxl
from io import BytesIO, TextIOWrapper

from kombu.message import Message as KombuMessage

from tider import Request
from tider.brokers import Broker
from tider.utils.log import get_logger
from tider.utils.misc import try_import
from tider.exceptions import ImproperlyConfigured

logger = get_logger(__name__)

rarfile = try_import('rarfile')


class Message(KombuMessage):

    def __init__(self, filename=None, raw=None, source=None, lineno=None, tell=None,
                 body=None, **kwargs):
        super().__init__(body=body, **kwargs)
        self.filename = filename
        self.tell = tell
        self.lineno = lineno

        self._raw = raw
        self._source = source

    @property
    def raw(self):
        return self._raw

    def ack(self, multiple=False):
        # maybe use file to store the acked lineno and position.
        self._state = 'ACK'

    def reject(self, requeue=False):
        self._state = 'REJECTED'

    def requeue(self):
        self._state = 'REQUEUED'


class FilesBroker(Broker):

    def _consume_json(self, file, on_message=None, on_message_consumed=None):
        lineno = 0
        filename = None
        if hasattr(file, 'name'):
            filename = file.name
        for line in file:
            lineno += 1
            message = Message(filename=filename, lineno=lineno, tell=file.tell(), body=line, content_type='application/json')
            on_message and on_message(message=message, payload=message.payload, no_ack=True)
            on_message_consumed and on_message_consumed()

    def _consume_xlsx(self, file, on_message=None, on_message_consumed=None):
        fp = BytesIO()
        zout = zipfile.ZipFile(fp, "w")
        zin = zipfile.ZipFile(file, "r")
        for item in zin.infolist():
            buffer = zin.read(item.filename)
            if item.filename == "xl/styles.xml":
                styles = buffer.decode("utf-8")
                styles = styles.replace("<x:fill />", "")
                # https://learn.microsoft.com/zh-cn/dotnet/api/documentformat.openxml.spreadsheet.fills?view=openxml-3.0.1
                styles = styles.replace('<fills count="1"><fill/></fills>', '').replace('<fill/>', '')
                cell_styles = re.findall(r'(<cellStyle\s*xfId.*?/>)', styles)
                for style in cell_styles:
                    if not re.findall(r'(name=".*?")', style):
                        new_style = style.replace("/>", ' name=""/>')
                        styles = styles.replace(style, new_style)
                buffer = styles.encode("utf-8")
            zout.writestr(item, buffer)
        zin.close()
        zout.close()

        wb = openpyxl.load_workbook(fp, read_only=True, data_only=True, keep_links=False, keep_vba=False)
        for name in wb.sheetnames:
            sheet = wb[name]
            titles = None
            for idx, row in enumerate(sheet.iter_rows(values_only=True)):
                if all([x is None for x in row]):
                    continue
                if idx == 0:
                    titles = row
                    continue
                message = dict(zip(titles, row))
                message.pop(None, None)
                on_message and on_message(message=message, payload=message, no_ack=True)
                on_message_consumed and on_message_consumed()
        fp.close()

    def _consume_xls(self, file, on_message=None, on_message_consumed=None):
        wb = xlrd.open_workbook(file_contents=file.read(), on_demand=True)
        for sheet in wb.sheets():
            invalid_count = 0
            titles = None
            for row in range(sheet.nrows):
                if row == 0:
                    titles = sheet.row_values(row)
                    continue
                if invalid_count > 100:
                    break
                if all([x is None or not str(x).strip() for x in sheet.row_values(row)]):
                    invalid_count += 1
                    continue
                message = dict(zip(titles, sheet.row_values(row)))
                message.pop(None, None)
                on_message and on_message(message=message, payload=message, no_ack=True)
                on_message_consumed and on_message_consumed()
                invalid_count = 0

    def _consume_csv(self, file, on_message=None, on_message_consumed=None):
        with TextIOWrapper(file, encoding='utf-8-sig') as csvfile:
            reader = csv.DictReader(csvfile)
            for message in reader:
                message.pop(None, None)
                on_message and on_message(message=message, payload=message, no_ack=True)
                on_message_consumed and on_message_consumed()

    def _consume_rar(self, file, on_message=None, on_message_consumed=None):
        if not rarfile:
            raise ImproperlyConfigured(
                'You need to install the rarfile library to consume messages from .rar files.'
            )
        with rarfile.RarFile(file) as rf:
            for f in rf.infolist():
                filename = f.filename
                with rf.open(filename) as crf:
                    ext = filename.rsplit('.', maxsplit=1)[-1]
                    if ext == 'json':
                        self._consume_json(crf, on_message, on_message_consumed)
                    elif ext == 'xlsx':
                        self._consume_xlsx(crf, on_message, on_message_consumed)
                    elif ext == 'xls':
                        self._consume_xls(crf, on_message, on_message_consumed)
                    elif ext == 'csv':
                        self._consume_csv(file, on_message, on_message_consumed)

    def _consume_zip(self, file, on_message=None, on_message_consumed=None):
        with zipfile.ZipFile(file) as zipf:
            for f in zipf.infolist():
                filename = f.filename
                with zipf.open(filename) as zrf:
                    ext = filename.rsplit('.', maxsplit=1)[-1]
                    if ext == 'json':
                        self._consume_json(zrf, on_message, on_message_consumed)
                    elif ext == 'xlsx':
                        self._consume_xlsx(zrf, on_message, on_message_consumed)
                    elif ext == 'xls':
                        self._consume_xls(zrf, on_message, on_message_consumed)
                    elif ext == 'csv':
                        self._consume_csv(file, on_message, on_message_consumed)

    def consume(self, queues=None, on_message=None, on_message_consumed=None):
        """Consume messages and return the connection to trigger consuming."""
        data_source = self.crawler.data_source
        if not data_source:
            logger.warning("No data source detected, please make sure you're using the right broker.")

        is_local_file = False
        file = None
        ext = data_source.rsplit('.', maxsplit=1)[-1]
        try:
            if not data_source.startswith(('http', 'ftp')):
                is_local_file = True
                file = open(data_source, mode='rb')
            elif data_source.startswith('http'):
                explorer = self.crawler.engine.explorer
                content = explorer.explore(Request(url=data_source)).content
                file = BytesIO(content)
            else:
                raise ValueError("Unsupported data source.")
            if ext == 'json':
                self._consume_json(file, on_message, on_message_consumed)
            elif ext == 'xlsx':
                self._consume_xlsx(file, on_message, on_message_consumed)
            elif ext == 'xls':
                self._consume_xls(file, on_message, on_message_consumed)
            elif ext == 'csv':
                self._consume_csv(file, on_message, on_message_consumed)
            elif ext == 'rar':
                self._consume_rar(file, on_message, on_message_consumed)
            elif ext == 'zip':
                self._consume_zip(file, on_message, on_message_consumed)
            else:
                message = {"raw": file, "source": data_source}
                on_message and on_message(message=message, payload=message, no_ack=True)
            if not self.crawler.engine._spider_closed.is_set():
                # start_requests consumed.
                self.crawler.engine._spider_closed.set()
            on_message_consumed and on_message_consumed(loop=True)
        except RuntimeError:
            return  # maybe shutdown
        finally:
            if is_local_file:
                file and file.close()
                if self.crawler.settings.getbool('BROKER_REMOVE_FILE_AFTER_CONSUMED', False):
                    os.remove(data_source)
